from datetime import date
import io
from flask import (Blueprint, render_template, request, jsonify, session, send_file)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from models import get_connection, registrar_log
from routes.auth import login_required
from .constantes import ESTADOS_ASISTENCIA, es_estado_valido

asistencia_bp = Blueprint(
    "asistencia", __name__,
    url_prefix="/asistencia"
)

def fichas_permitidas(cur):
    """Devuelve las fichas visibles para el usuario actual.
    Instructor -> solo sus fichas asignadas. superadmin/admin -> todas."""
    rol = session.get("admin_rol")
    if rol == "instructor":
        cur.execute(
            """SELECT f.id, f.numero
                 FROM fichas f
                 JOIN instructor_fichas ifi ON ifi.ficha_id = f.id
                WHERE ifi.instructor_id = %s
                ORDER BY f.numero""",
            (session.get("admin_id"),)
        )
    else:
        cur.execute("SELECT id, numero FROM fichas ORDER BY numero")
    return cur.fetchall()

def instructor_puede_ver_ficha(cur, ficha_id):
    if session.get("admin_rol") != "instructor":
        return True
    cur.execute(
        "SELECT 1 FROM instructor_fichas WHERE instructor_id = %s AND ficha_id = %s",
        (session.get("admin_id"), ficha_id)
    )
    return cur.fetchone() is not None

@asistencia_bp.route("/registrar", methods=["GET"])
@login_required
def registrar():
    db = get_connection(); cur = db.cursor(dictionary=True)
    fichas = fichas_permitidas(cur)

    ficha_id = request.args.get("ficha_id", type=int)
    fecha = request.args.get("fecha") or date.today().isoformat()

    aprendices, asistencia_previa = [], {}
    if ficha_id and instructor_puede_ver_ficha(cur, ficha_id):
        cur.execute(
            """SELECT id, tipo_documento, nombres, apellidos, identificacion
                 FROM usuarios WHERE ficha_id = %s ORDER BY apellidos, nombres""",
            (ficha_id,)
        )
        aprendices = cur.fetchall()
        cur.execute(
            "SELECT aprendiz_id, estado, observacion FROM asistencias WHERE ficha_id = %s AND fecha = %s",
            (ficha_id, fecha)
        )
        for row in cur.fetchall():
            asistencia_previa[row["aprendiz_id"]] = row

    cur.close(); db.close()
    return render_template(
        "asistencia/registrar.html",
        fichas=fichas, aprendices=aprendices, asistencia_previa=asistencia_previa,
        ficha_id=ficha_id, fecha=fecha, estados=ESTADOS_ASISTENCIA,
        active_tab='registrar'
    )

@asistencia_bp.route("/guardar", methods=["POST"])
@login_required
def guardar():
    data = request.get_json(silent=True) or {}
    ficha_id = data.get("ficha_id")
    fecha = data.get("fecha")
    registros = data.get("registros", [])
    if not ficha_id or not fecha:
        return jsonify({"ok": False, "error": "Falta ficha o fecha"}), 400

    db = get_connection(); cur = db.cursor()
    # Seguridad: el instructor solo puede guardar en sus fichas
    if not instructor_puede_ver_ficha(cur, ficha_id):
        cur.close(); db.close()
        return jsonify({"ok": False, "error": "No tienes esta ficha asignada"}), 403

    admin_id = session.get("admin_id"); guardados = 0
    try:
        for r in registros:
            estado = r.get("estado"); aprendiz_id = r.get("aprendiz_id")
            observacion = (r.get("observacion") or "")[:255]
            if not aprendiz_id or not es_estado_valido(estado):
                continue
            cur.execute(
                """INSERT INTO asistencias
                       (aprendiz_id, ficha_id, fecha, estado, observacion, registrado_por)
                   VALUES (%s, %s, %s, %s, %s, %s)
                   ON DUPLICATE KEY UPDATE
                       estado = VALUES(estado),
                       observacion = VALUES(observacion),
                       registrado_por = VALUES(registrado_por)""",
                (aprendiz_id, ficha_id, fecha, estado, observacion, admin_id)
            )
            guardados += 1
        db.commit()
        registrar_log(admin_id, session.get("admin_username"), "GUARDAR", "asistencia", ficha_id, f"fecha {fecha}: {guardados} registros")
        return jsonify({"ok": True, "guardados": guardados})
    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        cur.close(); db.close()

@asistencia_bp.route("/consultar", methods=["GET"])
@login_required
def consultar():
    db = get_connection(); cur = db.cursor(dictionary=True)
    fichas = fichas_permitidas(cur)
    ids_permitidos = {f["id"] for f in fichas}

    ficha_id = request.args.get("ficha_id", type=int)
    desde = request.args.get("desde"); hasta = request.args.get("hasta")
    estado = request.args.get("estado")

    where, params = ["1=1"], []
    # Restringe a las fichas permitidas del instructor
    if ids_permitidos:
        where.append(f"a.ficha_id IN ({','.join(['%s']*len(ids_permitidos))})")
        params.extend(ids_permitidos)
    else:
        where.append("1=0")  # instructor sin fichas -> sin resultados
    if ficha_id:
        where.append("a.ficha_id = %s"); params.append(ficha_id)
    if desde:
        where.append("a.fecha >= %s"); params.append(desde)
    if hasta:
        where.append("a.fecha <= %s"); params.append(hasta)
    if estado and es_estado_valido(estado):
        where.append("a.estado = %s"); params.append(estado)

    # Resumen de estados para todo el filtro
    cur.execute(
        f"""SELECT a.estado, COUNT(*) as total
             FROM asistencias a JOIN usuarios u ON u.id = a.aprendiz_id
            WHERE {' AND '.join(where)}
            GROUP BY a.estado""",
        params
    )
    resumen = {c: 0 for c in ESTADOS_ASISTENCIA}
    total_registros = 0
    for row in cur.fetchall():
        resumen[row["estado"]] = row["total"]
        total_registros += row["total"]

    # Paginación
    page = request.args.get("page", 1, type=int)
    per_page = 20
    offset = (page - 1) * per_page
    total_pages = (total_registros + per_page - 1) // per_page if total_registros > 0 else 1

    cur.execute(
        f"""SELECT a.fecha, a.estado, a.observacion, u.nombres, u.apellidos, u.identificacion, u.tipo_documento
             FROM asistencias a JOIN usuarios u ON u.id = a.aprendiz_id
            WHERE {' AND '.join(where)}
            ORDER BY a.fecha DESC, u.apellidos
            LIMIT %s OFFSET %s""",
        params + [per_page, offset]
    )
    filas = cur.fetchall()
    cur.close(); db.close()

    return render_template(
        "asistencia/consultar.html",
        fichas=fichas, filas=filas, resumen=resumen, estados=ESTADOS_ASISTENCIA,
        ficha_id=ficha_id, desde=desde, hasta=hasta, estado=estado,
        page=page, total_pages=total_pages,
        active_tab='consultar'
    )

@asistencia_bp.route("/api/resumen", methods=["GET"])
@login_required
def api_resumen():
    db = get_connection(); cur = db.cursor(dictionary=True)
    ficha_id = request.args.get("ficha_id", type=int)
    
    where = ""
    params = ()
    if ficha_id:
        if instructor_puede_ver_ficha(cur, ficha_id):
            where = "WHERE ficha_id = %s"
            params = (ficha_id,)
        else:
            where = "WHERE 1=0"
    else:
        if session.get("admin_rol") == "instructor":
            fichas = fichas_permitidas(cur)
            fids = [f["id"] for f in fichas]
            if fids:
                format_strings = ','.join(['%s'] * len(fids))
                where = f"WHERE ficha_id IN ({format_strings})"
                params = tuple(fids)
            else:
                where = "WHERE 1=0"

    cur.execute(f"SELECT estado, COUNT(*) AS total FROM asistencias {where} GROUP BY estado", params)
    datos = {c: 0 for c in ESTADOS_ASISTENCIA}
    for row in cur.fetchall():
        if row["estado"] in datos:
            datos[row["estado"]] = row["total"]
    cur.close(); db.close()
    return jsonify({
        "labels": [ESTADOS_ASISTENCIA[c]["label"] for c in datos],
        "data":   [datos[c] for c in datos],
        "colors": [ESTADOS_ASISTENCIA[c]["color"] for c in datos],
    })


@asistencia_bp.route("/informe", methods=["GET"])
@login_required
def informe():
    db = get_connection(); cur = db.cursor(dictionary=True)
    fichas = fichas_permitidas(cur)

    ficha_id = request.args.get("ficha_id", type=int)
    aprendiz_id = request.args.get("aprendiz_id", type=int)
    desde = request.args.get("desde"); hasta = request.args.get("hasta")

    aprendices_lista = []
    datos = []
    detalle_dias = []
    numero_ficha = ""
    nombre_programa = ""
    total_dias = 0

    if ficha_id and instructor_puede_ver_ficha(cur, ficha_id):
        # Info de la ficha
        cur.execute(
            """SELECT f.numero, p.nombre AS nombre_programa
                 FROM fichas f LEFT JOIN programas p ON p.id = f.programa_id
                WHERE f.id = %s""", (ficha_id,))
        info = cur.fetchone() or {}
        numero_ficha = info.get("numero", "")
        nombre_programa = info.get("nombre_programa", "")

        # Lista de aprendices de esta ficha (para el select)
        cur.execute(
            """SELECT id, nombres, apellidos, identificacion
                 FROM usuarios WHERE ficha_id = %s ORDER BY apellidos, nombres""",
            (ficha_id,))
        aprendices_lista = cur.fetchall()

        # Construir filtro
        where = ["a.ficha_id = %s"]; params = [ficha_id]
        if desde:
            where.append("a.fecha >= %s"); params.append(desde)
        if hasta:
            where.append("a.fecha <= %s"); params.append(hasta)
        if aprendiz_id:
            where.append("a.aprendiz_id = %s"); params.append(aprendiz_id)

        # Total días únicos
        cur.execute(
            f"SELECT COUNT(DISTINCT a.fecha) AS total FROM asistencias a WHERE {' AND '.join(where)}",
            params)
        total_dias = (cur.fetchone() or {}).get("total", 0)

        # Resumen por aprendiz
        cur.execute(
            f"""SELECT u.id AS aprendiz_id, u.nombres, u.apellidos, u.identificacion,
                       SUM(a.estado = 'A') AS asiste,
                       SUM(a.estado = 'CE') AS con_excusa,
                       SUM(a.estado = 'SE') AS sin_excusa,
                       SUM(a.estado = 'INC') AS incapacidad,
                       SUM(a.estado = 'LIC') AS licencia,
                       SUM(a.estado = 'CLM') AS calamidad,
                       SUM(a.estado = 'SFF') AS sin_formacion,
                       SUM(a.estado = 'RET') AS retirado,
                       COUNT(*) AS total_registros
                  FROM asistencias a
                  JOIN usuarios u ON u.id = a.aprendiz_id
                 WHERE {' AND '.join(where)}
                 GROUP BY u.id, u.nombres, u.apellidos, u.identificacion
                 ORDER BY u.apellidos, u.nombres""",
            params)
        for row in cur.fetchall():
            total = row["total_registros"] or 1
            row["pct_asistencia"] = round((row["asiste"] / total) * 100)
            datos.append(row)

        # Si se pidió detalle de un aprendiz individual
        if aprendiz_id:
            cur.execute(
                f"""SELECT a.fecha, a.estado, a.observacion
                      FROM asistencias a
                     WHERE {' AND '.join(where)}
                     ORDER BY a.fecha""",
                params)
            detalle_dias = cur.fetchall()

    cur.close(); db.close()
    return render_template(
        "asistencia/informe.html",
        fichas=fichas, ficha_id=ficha_id, aprendiz_id=aprendiz_id,
        desde=desde, hasta=hasta, datos=datos, detalle_dias=detalle_dias,
        aprendices_lista=aprendices_lista, numero_ficha=numero_ficha,
        nombre_programa=nombre_programa, total_dias=total_dias,
        estados=ESTADOS_ASISTENCIA, active_tab='informe'
    )


@asistencia_bp.route("/api/aprendices_ficha", methods=["GET"])
@login_required
def api_aprendices_ficha():
    ficha_id = request.args.get("ficha_id", type=int)
    if not ficha_id:
        return jsonify([])
    db = get_connection(); cur = db.cursor(dictionary=True)
    if not instructor_puede_ver_ficha(cur, ficha_id):
        cur.close(); db.close()
        return jsonify([])
    cur.execute(
        """SELECT id, nombres, apellidos FROM usuarios
             WHERE ficha_id = %s ORDER BY apellidos, nombres""",
        (ficha_id,))
    resultado = cur.fetchall()
    cur.close(); db.close()
    return jsonify(resultado)


@asistencia_bp.route("/exportar/excel", methods=["GET"])
@login_required
def exportar_excel():
    result = _generar_excel_asistencia()
    if isinstance(result[0], str):  # Error response (message, status_code)
        return result
    wb, nombre_archivo = result
    import io
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"{nombre_archivo}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@asistencia_bp.route("/exportar/pdf", methods=["GET"])
@login_required
def exportar_pdf():
    result = _generar_excel_asistencia()
    if isinstance(result[0], str):  # Error response (message, status_code)
        return result
    wb, nombre_archivo = result
    import os, tempfile, subprocess, sys
    from flask import current_app
    tmp_dir = tempfile.mkdtemp(dir=current_app.config['GENERADOS_FOLDER'])
    xlsx_path = os.path.join(tmp_dir, f"{nombre_archivo}.xlsx")
    pdf_path = os.path.join(tmp_dir, f"{nombre_archivo}.pdf")
    wb.save(xlsx_path)
    
    script = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'convert_to_pdf.py')
    proc = subprocess.run([sys.executable, script, xlsx_path, pdf_path], capture_output=True)
    
    if os.path.exists(pdf_path):
        with open(pdf_path, 'rb') as f:
            pdf_data = f.read()
        import shutil
        shutil.rmtree(tmp_dir, ignore_errors=True)
        import io
        return send_file(io.BytesIO(pdf_data), as_attachment=True,
                         download_name=f"{nombre_archivo}.pdf", mimetype="application/pdf")
    else:
        import shutil
        shutil.rmtree(tmp_dir, ignore_errors=True)
        current_app.logger.error("Error al generar PDF. STDOUT: %s, STDERR: %s", proc.stdout, proc.stderr)
        return "Error al generar PDF", 500

def _generar_excel_asistencia():
    from flask import request
    from datetime import datetime

    ficha_id = request.args.get("ficha_id", type=int)
    desde = request.args.get("desde"); hasta = request.args.get("hasta")
    db = get_connection(); cur = db.cursor(dictionary=True)
    if not instructor_puede_ver_ficha(cur, ficha_id):
        cur.close(); db.close()
        return "No autorizado", 403

    # Obtener información de la ficha y programa para el encabezado
    cur.execute(
        """SELECT f.numero AS numero_ficha, p.nombre AS nombre_programa
             FROM fichas f
             LEFT JOIN programas p ON p.id = f.programa_id
            WHERE f.id = %s""",
        (ficha_id,)
    )
    info_ficha = cur.fetchone() or {}
    numero_ficha = info_ficha.get("numero_ficha", "")
    nombre_programa = info_ficha.get("nombre_programa", "")

    where = ["a.ficha_id = %s"]; params = [ficha_id]
    if desde:
        where.append("a.fecha >= %s"); params.append(desde)
    if hasta:
        where.append("a.fecha <= %s"); params.append(hasta)

    # Incluir observacion en la consulta
    cur.execute(
        f"""SELECT u.tipo_documento, u.identificacion, u.nombres, u.apellidos,
                   a.fecha, a.estado, a.observacion
             FROM asistencias a JOIN usuarios u ON u.id = a.aprendiz_id
            WHERE {' AND '.join(where)}
            ORDER BY u.apellidos, u.nombres, a.fecha""",
        params
    )
    filas = cur.fetchall(); cur.close(); db.close()

    fechas = sorted({f["fecha"] for f in filas})
    aprendices = {}
    observaciones_aprendiz = {}
    for f in filas:
        key = (f["tipo_documento"], f["identificacion"], f["nombres"], f["apellidos"])
        aprendices.setdefault(key, {})[f["fecha"]] = f["estado"]
        # Recopilar observaciones únicas del aprendiz
        obs = (f.get("observacion") or "").strip()
        if obs:
            observaciones_aprendiz.setdefault(key, set()).add(obs)

    wb = Workbook(); ws = wb.active; ws.title = "Asistencia"

    # ─── ESTILOS ────────────────────────────────────────────────
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    titulo_font = Font(bold=True, size=16, color="FFFFFF")
    subtitulo_font = Font(bold=True, size=11, color="305496")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    titulo_fill = PatternFill("solid", fgColor="305496")

    # ─── CALCULAR TOTAL DE COLUMNAS ────────────────────────────
    columnas_estados = ["A", "CE", "SE", "INC", "LIC", "CLM", "SFF", "RET"]

    header = (["Tipo Doc", "Identificación", "Nombres", "Apellidos"]
              + [d.isoformat() if hasattr(d, 'isoformat') else str(d) for d in fechas]
              + columnas_estados
              + ["% INASIST.", "Observación"])

    total_cols = len(header)

    # ─── ENCABEZADO INSTITUCIONAL ───────────────────────────────
    # Fila 1: Título principal
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    celda_titulo = ws.cell(row=1, column=1, value="LISTA DE ASISTENCIA")
    celda_titulo.font = titulo_font
    celda_titulo.fill = titulo_fill
    celda_titulo.alignment = Alignment(horizontal="center", vertical="center")

    # Fila 2: Ficha y Programa
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    celda_ficha = ws.cell(row=2, column=1,
                          value=f"Ficha: {numero_ficha}   |   Programa: {nombre_programa}")
    celda_ficha.font = subtitulo_font
    celda_ficha.alignment = Alignment(horizontal="center", vertical="center")

    # Fila 3: Rango de fechas del reporte
    texto_desde = desde if desde else "Inicio"
    texto_hasta = hasta if hasta else "Actual"
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)
    celda_rango = ws.cell(row=3, column=1,
                          value=f"Período del reporte:  Desde {texto_desde}  —  Hasta {texto_hasta}")
    celda_rango.font = subtitulo_font
    celda_rango.alignment = Alignment(horizontal="center", vertical="center")

    # Fila 4: Fecha y hora de impresión
    ahora = datetime.now().strftime("%Y-%m-%d  %H:%M:%S")
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=total_cols)
    celda_impresion = ws.cell(row=4, column=1,
                              value=f"Fecha de impresión: {ahora}")
    celda_impresion.font = Font(italic=True, size=10, color="555555")
    celda_impresion.alignment = Alignment(horizontal="center", vertical="center")

    # Fila 5: Registrado por (nombre del instructor actual)
    nombre_instructor = session.get("admin_nombre", "")
    rol_instructor = session.get("admin_rol", "")
    texto_registrado = f"Registrado por: {nombre_instructor}"
    if rol_instructor:
        texto_registrado += f"  ({rol_instructor})"
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=total_cols)
    celda_instructor = ws.cell(row=5, column=1, value=texto_registrado)
    celda_instructor.font = Font(bold=True, size=11, color="1B4332")
    celda_instructor.alignment = Alignment(horizontal="center", vertical="center")

    # Fila 6: vacía (separador)
    fila_inicio_datos = 7  # La fila donde empiezan los encabezados de columna

    # ─── ENCABEZADOS DE COLUMNA ─────────────────────────────────
    for col_idx, titulo in enumerate(header, start=1):
        celda = ws.cell(row=fila_inicio_datos, column=col_idx, value=titulo)
        celda.font = header_font
        celda.fill = header_fill
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = thin_border

    # ─── DATOS DE APRENDICES ────────────────────────────────────
    fila_actual = fila_inicio_datos + 1
    for (tdoc, ident, nom, ape), reg in aprendices.items():
        datos_fila = [tdoc or "", ident or "", nom or "", ape or ""]

        # Asistencias por fecha
        for d in fechas:
            datos_fila.append(reg.get(d, ""))

        # Conteos por estado
        conteo_est = {k: 0 for k in columnas_estados}
        for estado in reg.values():
            if estado in conteo_est:
                conteo_est[estado] += 1

        for col in columnas_estados:
            datos_fila.append(conteo_est[col])

        # Calcular % inasistencias
        total_sesiones = len(reg)
        total_inasistencias = sum(conteo_est[k] for k in ["CE", "SE", "INC", "LIC", "CLM"])
        porcentaje = (total_inasistencias / total_sesiones) if total_sesiones > 0 else 0
        datos_fila.append(f"{porcentaje:.0%}")

        # Observación: unir todas las observaciones únicas del aprendiz
        obs_set = observaciones_aprendiz.get((tdoc, ident, nom, ape), set())
        datos_fila.append("; ".join(sorted(obs_set)))

        # Escribir la fila
        for col_idx, valor in enumerate(datos_fila, start=1):
            celda = ws.cell(row=fila_actual, column=col_idx, value=valor)
            celda.border = thin_border
            if col_idx == total_cols:
                celda.alignment = Alignment(wrap_text=True, vertical="center")
            else:
                celda.alignment = Alignment(vertical="center")

        # Aplicar colores a las celdas de estado por fecha
        for i, d in enumerate(fechas):
            estado = reg.get(d, "")
            if estado in ESTADOS_ASISTENCIA:
                color_hex = "FF" + ESTADOS_ASISTENCIA[estado]["color"].lstrip("#")
                celda = ws.cell(row=fila_actual, column=5 + i)
                celda.fill = PatternFill("solid", fgColor=color_hex)
                celda.font = Font(color="FFFFFF", bold=True)
                celda.alignment = Alignment(horizontal="center", vertical="center")

        fila_actual += 1

    # ─── AJUSTAR ANCHO DE COLUMNAS ──────────────────────────────
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        max_length = 0
        # Obtener la letra de columna de forma segura (las MergedCell no tienen column_letter)
        first_cell = col[0]
        if isinstance(first_cell, MergedCell):
            continue
        col_letter = get_column_letter(first_cell.column)
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # Limitar ancho máximo para columnas muy anchas (como Observación)
        adjusted_width = min(max_length + 2, 40)
        # Si es la última columna (Observación), forzar un ancho razonable para que el wrap funcione bien
        if col_letter == get_column_letter(total_cols):
            adjusted_width = 40
            
        ws.column_dimensions[col_letter].width = max(adjusted_width, 10)

    # ─── HOJA RESUMEN ──────────────────────────────────────────
    resumen = wb.create_sheet("Resumen"); conteo = {}
    for f in filas: conteo[f["estado"]] = conteo.get(f["estado"], 0) + 1
    resumen.append(["Estado", "Total"])
    for cod, total in conteo.items(): resumen.append([cod, total])

    # ─── CONFIGURACIÓN DE IMPRESIÓN (para PDF) ─────────────────
    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # tantas páginas como necesite verticalmente
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.3, right=0.3, top=0.4, bottom=0.4,
        header=0.2, footer=0.2
    )
    ws.print_title_rows = f'1:{fila_inicio_datos}'  # repetir encabezado en cada página

    # ─── NOMBRE DESCRIPTIVO DEL ARCHIVO ────────────────────────
    nombre_archivo = f"asistencia_ficha_{numero_ficha}"
    if desde:
        nombre_archivo += f"_{desde}"
    if hasta:
        nombre_archivo += f"_{hasta}"

    return wb, nombre_archivo
