import os
import uuid
import shutil
import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Mapeo de celdas para aprendices (1 a 5)
APRENDIZ_MAP = {
    1: {'nombre': 'B15', 'doc': 'E15', 'tel': 'H15', 'correo_inst': 'B21', 'correo_pers': 'E21', 'dir': 'H21',
        'arl_nombre': 'B71', 'arl_afil': 'F71', 'arl_riesgo': 'G71', 'arl_corr': 'H71', 'arl_epp': 'I71',
        'firma': 'C79'},
    2: {'nombre': 'B16', 'doc': 'E16', 'tel': 'H16', 'correo_inst': 'B22', 'correo_pers': 'E22', 'dir': 'H22',
        'arl_nombre': 'B72', 'arl_afil': 'F72', 'arl_riesgo': 'G72', 'arl_corr': 'H72', 'arl_epp': 'I72',
        'firma': 'H79'},
    3: {'nombre': 'B17', 'doc': 'E17', 'tel': 'H17', 'correo_inst': 'B23', 'correo_pers': 'E23', 'dir': 'H23',
        'arl_nombre': 'B73', 'arl_afil': 'F73', 'arl_riesgo': 'G73', 'arl_corr': 'H73', 'arl_epp': 'I73',
        'firma': 'C82'},
    4: {'nombre': 'B18', 'doc': 'E18', 'tel': 'H18', 'correo_inst': 'B24', 'correo_pers': 'E24', 'dir': 'H24',
        'arl_nombre': 'B74', 'arl_afil': 'F74', 'arl_riesgo': 'G74', 'arl_corr': 'H74', 'arl_epp': 'I74',
        'firma': 'H82'},
    5: {'nombre': 'B19', 'doc': 'E19', 'tel': 'H19', 'correo_inst': 'B25', 'correo_pers': 'E25', 'dir': 'H25',
        'arl_nombre': 'B75', 'arl_afil': 'F75', 'arl_riesgo': 'G75', 'arl_corr': 'H75', 'arl_epp': 'I75',
        'firma': 'C85'}
}

# Mapeo de Alternativas
ALTERNATIVA_MAP = {
    'Contrato aprendizaje': 'C43',
    'Contrato vínculo formativo': 'C47',
    'Monitoria': 'H43',
    'Proyecto productivo': 'H45',
    'Vínculo laboral': 'H47'
}

# Filas base para actividades
ACTIVIDADES_ROWS = [53, 55, 57, 59, 61]

def parse_date(date_str):
    """Convierte un string YYYY-MM-DD o YYYY/MM/DD a datetime.date."""
    if not date_str:
        return None
    if isinstance(date_str, datetime.date):
        return date_str
    try:
        return datetime.datetime.strptime(str(date_str).replace('/', '-').split('T')[0], '%Y-%m-%d').date()
    except ValueError:
        return None

def formato_fecha(fecha):
    """Formatea la fecha para el string Desde... Hasta..."""
    if not fecha:
        return ""
    if isinstance(fecha, str):
        fecha = parse_date(fecha)
    return fecha.strftime("%d/%m/%Y")

def agregar_firma(ws, path, celda_ancla):
    if not path or not os.path.exists(path):
        return
    try:
        img = Image(path)
        img.width = 130
        img.height = 35
        ws.add_image(img, celda_ancla)
    except Exception as e:
        print(f"Error al agregar imagen de firma {path}: {e}")

def generar_excel_bitacora(template_path, data, output_path=None):
    """
    Rellena la plantilla de Bitácora Articulación con la Media con los datos proporcionados.
    `data` debe ser un diccionario con la estructura esperada.
    Retorna la ruta del archivo .xlsx generado temporalmente.
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"No se encontró la plantilla base: {template_path}")

    # Copiar plantilla a archivo temporal
    if not output_path:
        tmp_dir = os.path.join('static', 'uploads', 'bam_tmp')
        os.makedirs(tmp_dir, exist_ok=True)
        filename = f"bitacora_{uuid.uuid4().hex}.xlsx"
        output_path = os.path.join(tmp_dir, filename)

    shutil.copy2(template_path, output_path)

    wb = load_workbook(output_path, data_only=False)
    
    if 'Formato Bitácora Art. Media' not in wb.sheetnames:
        raise ValueError('La plantilla no contiene la hoja "Formato Bitácora Art. Media"')
        
    ws = wb['Formato Bitácora Art. Media']

    # 1. Encabezado
    ws['B11'] = data.get('numero_bitacora', '')
    fecha_desde = formato_fecha(data.get('periodo_desde'))
    fecha_hasta = formato_fecha(data.get('periodo_hasta'))
    if fecha_desde or fecha_hasta:
        ws['E11'] = f"Desde {fecha_desde} hasta {fecha_hasta}"

    # 2. Aprendices
    from openpyxl.styles import Alignment
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    aprendices = data.get('aprendices', [])
    for idx, apr in enumerate(aprendices):
        orden = idx + 1
        if orden > 5:
            break
            
        mapa = APRENDIZ_MAP[orden]
        
        ws[mapa['nombre']] = f"{apr.get('nombres', '')} {apr.get('apellidos', '')}".strip()
        ws[mapa['doc']] = f"{apr.get('tipo_documento', '')} {apr.get('identificacion', '')}".strip()
        ws[mapa['tel']] = apr.get('telefono', '')
        ws[mapa['correo_inst']] = apr.get('correo_institucional', '')
        ws[mapa['correo_pers']] = apr.get('correo_personal', '')
        ws[mapa['dir']] = apr.get('direccion', '')

        # Centrar datos del aprendiz
        ws[mapa['nombre']].alignment = center_alignment
        ws[mapa['doc']].alignment = center_alignment
        ws[mapa['tel']].alignment = center_alignment
        ws[mapa['correo_inst']].alignment = center_alignment
        ws[mapa['correo_pers']].alignment = center_alignment
        ws[mapa['dir']].alignment = center_alignment

        ws[mapa['arl_nombre']] = f"{apr.get('nombres', '')} {apr.get('apellidos', '')}".strip()
        ws[mapa['arl_afil']] = apr.get('arl_afiliado', 'SI')
        ws[mapa['arl_riesgo']] = apr.get('arl_nivel_riesgo', 1)
        ws[mapa['arl_corr']] = apr.get('arl_corresponde', 'SI')
        ws[mapa['arl_epp']] = apr.get('arl_epp', 'SI')
        
        # Firma aprendiz
        firma_path = apr.get('firma_path')
        if firma_path:
            agregar_firma(ws, firma_path, mapa['firma'])

    # Si hay menos de 5 aprendices, vaciamos las celdas de los restantes
    for idx in range(len(aprendices), 5):
        orden = idx + 1
        mapa = APRENDIZ_MAP[orden]
        ws[mapa['nombre']] = ""
        ws[mapa['doc']] = ""
        ws[mapa['tel']] = ""
        ws[mapa['correo_inst']] = ""
        ws[mapa['correo_pers']] = ""
        ws[mapa['dir']] = ""


    # 3. Datos del grupo
    ws['B27'] = data.get('ficha_numero', '')
    ws['C27'] = data.get('modalidad_formacion', '')
    ws['F27'] = data.get('programa_nombre', '')
    ws['I27'] = data.get('modalidad_ejecucion', '')

    # 4. Ente co-formador
    ws['B31'] = data.get('entidad_nombre', '')
    ws['F31'] = data.get('entidad_nit', '')
    ws['H31'] = data.get('entidad_direccion', '')

    ws['B35'] = data.get('jefe_nombre', '')
    ws['D35'] = data.get('jefe_cargo', '')
    ws['F35'] = data.get('jefe_telefono', '')
    ws['H35'] = data.get('jefe_correo', '')

    # 5. Instructor de Seguimiento
    ws['B39'] = data.get('seguimiento_nombre', '')
    ws['G39'] = data.get('seguimiento_correo', '')

    # 6. Alternativa Etapa Productiva
    alt = data.get('alternativa_etapa')
    if alt in ALTERNATIVA_MAP:
        ws[ALTERNATIVA_MAP[alt]] = 'X'

    # 7. Actividades
    actividades = data.get('actividades', [])
    for idx in range(5):
        r = ACTIVIDADES_ROWS[idx]
        if idx < len(actividades):
            act = actividades[idx]
            # Escribimos los datos de la actividad
            ws[f'B{r}'] = act.get('descripcion', '')
            
            # Competencias se sobrescribe SOLO si el usuario lo envia explícitamente y no esta vacio
            comp = act.get('competencias')
            if comp:
                ws[f'D{r}'] = comp
            else:
                ws[f'D{r}'] = ''
                
            ws[f'F{r}'] = parse_date(act.get('fecha_inicio'))
            ws[f'G{r}'] = parse_date(act.get('fecha_fin'))
            
            # Formato de celda para fechas dd/mm/aa
            ws[f'F{r}'].number_format = 'dd/mm/yy'
            ws[f'G{r}'].number_format = 'dd/mm/yy'
            
            ws[f'H{r}'] = act.get('evidencia', '')
            ws[f'I{r}'] = act.get('observaciones', '')
            
            # Centrar las observaciones
            ws[f'B{r}'].alignment = center_alignment
            ws[f'D{r}'].alignment = center_alignment
            ws[f'H{r}'].alignment = center_alignment
            ws[f'I{r}'].alignment = center_alignment
        else:
            # Vaciar filas de actividades sobrantes si vienen con texto de la plantilla
            ws[f'B{r}'] = ''
            ws[f'D{r}'] = ''
            ws[f'F{r}'] = ''
            ws[f'G{r}'] = ''
            ws[f'H{r}'] = ''
            ws[f'I{r}'] = ''

    # 8. Fecha Entrega y Firma Ente Co-formador
    ws['H86'] = parse_date(data.get('fecha_entrega'))
    ws['H86'].number_format = 'dd/mm/yy'

    # La firma del instructor de seguimiento en C92 se deja en blanco según requisitos.
    # Firma ente co-formador (instructor logueado)
    firma_jefe = data.get('firma_ente_coformador_path')
    if firma_jefe:
        agregar_firma(ws, firma_jefe, 'H91')

    wb.save(output_path)
    wb.close()
    
    return output_path
